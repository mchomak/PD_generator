"""Excel file reading and validation for PD Generator."""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from config import ColumnMapping

logger = logging.getLogger(__name__)


@dataclass
class ProjectData:
    """Data for a single project."""
    project_id: str
    project_name: str
    problem: str
    solution: str
    product: str
    team: str
    image_filename: Optional[str] = None
    row_number: int = 0  # For error reporting

    def validate(self) -> List[str]:
        """Validate project data, return list of errors."""
        errors = []

        if not self.project_id:
            errors.append("Missing project_id")
        if not self.project_name:
            errors.append("Missing project_name")
        if not self.problem:
            errors.append("Missing problem description")
        if not self.solution:
            errors.append("Missing solution description")
        if not self.product:
            errors.append("Missing product description")
        if not self.team:
            errors.append("Missing team information")

        return errors


class ExcelReader:
    """Reader for project data from Excel files."""

    def __init__(self, file_path: Path, column_mapping: ColumnMapping):
        """
        Initialize Excel reader.

        Args:
            file_path: Path to the Excel file
            column_mapping: Mapping of column names to project fields
        """
        self.file_path = file_path
        self.column_mapping = column_mapping
        self._column_indices: Dict[str, int] = {}

    def _find_columns(self, header_row: Tuple[Any, ...]) -> Dict[str, int]:
        """Find column indices based on header row."""
        indices = {}
        headers = [str(cell).strip().lower() if cell else "" for cell in header_row]

        # Map each required field to its column index
        field_mappings = {
            "project_id": self.column_mapping.project_id,
            "project_name": self.column_mapping.project_name,
            "problem": self.column_mapping.problem,
            "solution": self.column_mapping.solution,
            "product": self.column_mapping.product,
            "team": self.column_mapping.team,
        }

        # Add optional image_filename mapping
        if self.column_mapping.image_filename:
            field_mappings["image_filename"] = self.column_mapping.image_filename

        for field, column_name in field_mappings.items():
            column_name_lower = column_name.lower().strip()
            try:
                idx = headers.index(column_name_lower)
                indices[field] = idx
            except ValueError:
                if field != "image_filename":  # image_filename is optional
                    logger.warning(f"Column '{column_name}' not found in Excel headers")

        return indices

    def _get_cell_value(self, row: Tuple[Any, ...], field: str) -> str:
        """Get cell value for a field, handling None and type conversion."""
        if field not in self._column_indices:
            return ""

        idx = self._column_indices[field]
        if idx >= len(row):
            return ""

        value = row[idx]
        if value is None:
            return ""

        # Convert to string and strip whitespace
        return str(value).strip()

    def read_projects(self) -> Tuple[List[ProjectData], List[str]]:
        """
        Read all projects from the Excel file.

        Returns:
            Tuple of (list of ProjectData, list of global errors)
        """
        global_errors = []

        if not self.file_path.exists():
            global_errors.append(f"Excel file not found: {self.file_path}")
            return [], global_errors

        try:
            workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        except Exception as e:
            global_errors.append(f"Failed to open Excel file: {e}")
            return [], global_errors

        # Use the first sheet
        sheet = workbook.active
        if sheet is None:
            global_errors.append("Excel file has no active sheet")
            return [], global_errors

        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            global_errors.append("Excel file must have at least a header row and one data row")
            return [], global_errors

        # Find column indices from header row
        self._column_indices = self._find_columns(rows[0])

        # Check for required columns
        required_fields = ["project_id", "project_name", "problem", "solution", "product", "team"]
        missing_columns = [f for f in required_fields if f not in self._column_indices]
        if missing_columns:
            global_errors.append(f"Missing required columns: {', '.join(missing_columns)}")
            return [], global_errors

        # Read data rows
        projects = []
        for row_num, row in enumerate(rows[1:], start=2):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            project = ProjectData(
                project_id=self._get_cell_value(row, "project_id"),
                project_name=self._get_cell_value(row, "project_name"),
                problem=self._get_cell_value(row, "problem"),
                solution=self._get_cell_value(row, "solution"),
                product=self._get_cell_value(row, "product"),
                team=self._get_cell_value(row, "team"),
                image_filename=self._get_cell_value(row, "image_filename") or None,
                row_number=row_num,
            )

            projects.append(project)

        workbook.close()

        logger.info(f"Read {len(projects)} projects from Excel file")
        return projects, global_errors


def find_project_image(
    project: ProjectData,
    images_folder: Path,
    supported_extensions: List[str] = None,
) -> Optional[Path]:
    """
    Find image file for a project.

    Args:
        project: Project data
        images_folder: Folder containing project images
        supported_extensions: List of supported image extensions

    Returns:
        Path to image file, or None if not found
    """
    if supported_extensions is None:
        supported_extensions = [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff"]

    # First, check if image_filename is specified
    if project.image_filename:
        image_path = images_folder / project.image_filename
        if image_path.exists():
            return image_path
        # Also try with the filename as-is (might be an absolute path)
        direct_path = Path(project.image_filename)
        if direct_path.exists():
            return direct_path

    # Otherwise, search by project_id
    for ext in supported_extensions:
        image_path = images_folder / f"{project.project_id}{ext}"
        if image_path.exists():
            return image_path

        # Try lowercase extension
        image_path = images_folder / f"{project.project_id}{ext.lower()}"
        if image_path.exists():
            return image_path

        # Try uppercase extension
        image_path = images_folder / f"{project.project_id}{ext.upper()}"
        if image_path.exists():
            return image_path

    return None
