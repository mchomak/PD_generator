#!/usr/bin/env python3
"""Script to create an example Excel file for PD Generator."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Projects"

# Define headers
headers = [
    "project_id",
    "project_name",
    "problem",
    "solution",
    "product",
    "team",
    "image_filename",
]

# Style for headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Example projects with Cyrillic support demonstration
example_projects = [
    {
        "project_id": "101",
        "project_name": "Smart Campus Navigation",
        "problem": "Students and visitors often struggle to find specific classrooms, labs, and offices within the large university campus. Traditional maps are outdated and don't provide real-time information about room availability or the fastest routes.",
        "solution": "We developed a mobile application that uses indoor positioning technology and real-time data integration to provide turn-by-turn navigation within campus buildings. The app syncs with the university's scheduling system to show room availability and suggests optimal routes based on current foot traffic.",
        "product": "SmartCampus Navigator - A cross-platform mobile application available on iOS and Android. Features include AR-based navigation, accessibility routes, integration with class schedules, and real-time updates on building access.",
        "team": "Alice Johnson (PM)\nBob Smith (Lead Dev)\nCarol Williams (UX Designer)\nDavid Brown (Backend Dev)",
        "image_filename": "",
    },
    {
        "project_id": "102",
        "project_name": "Умный мониторинг энергии",  # Russian: Smart Energy Monitoring
        "problem": "Университетские здания потребляют значительное количество энергии, причём большая часть тратится впустую из-за неэффективного использования. Отсутствие детального мониторинга затрудняет выявление и устранение источников потерь энергии.",
        "solution": "Мы создали комплексную систему мониторинга энергопотребления с использованием IoT-датчиков и машинного обучения для анализа паттернов использования. Система предоставляет рекомендации по оптимизации и автоматически управляет освещением и климатом.",
        "product": "EnergyWatch Dashboard - Веб-платформа для мониторинга энергопотребления в реальном времени. Включает прогнозирование потребления, автоматические оповещения и интеграцию с системами управления зданием.",
        "team": "Иван Петров (Руководитель)\nМария Сидорова (Разработчик)\nАлексей Козлов (Data Science)\nЕлена Новикова (IoT Engineer)",
        "image_filename": "",
    },
    {
        "project_id": "103",
        "project_name": "Lab Equipment Scheduler",
        "problem": "Research laboratories have expensive equipment that is often underutilized due to poor scheduling. Researchers waste time trying to book equipment, and conflicts arise when multiple teams need the same resources.",
        "solution": "Our solution is a centralized booking platform with an intelligent scheduling algorithm that maximizes equipment utilization while respecting priority levels and maintenance windows. The system integrates with existing lab management software.",
        "product": "LabBook Pro - Web-based scheduling platform with mobile companion app. Features include automated conflict resolution, usage analytics, maintenance tracking, and integration with research grant systems.",
        "team": "Dr. Emily Chen (PI)\nMichael Lee (Full Stack Dev)\nSarah Park (Product Manager)\nJames Wilson (QA Engineer)",
        "image_filename": "",
    },
]

# Write data rows
for row_num, project in enumerate(example_projects, 2):
    ws.cell(row=row_num, column=1, value=project["project_id"])
    ws.cell(row=row_num, column=2, value=project["project_name"])
    ws.cell(row=row_num, column=3, value=project["problem"])
    ws.cell(row=row_num, column=4, value=project["solution"])
    ws.cell(row=row_num, column=5, value=project["product"])
    ws.cell(row=row_num, column=6, value=project["team"])
    ws.cell(row=row_num, column=7, value=project["image_filename"])

    # Add borders
    for col in range(1, 8):
        ws.cell(row=row_num, column=col).border = thin_border

# Adjust column widths
column_widths = {
    "A": 12,  # project_id
    "B": 30,  # project_name
    "C": 50,  # problem
    "D": 50,  # solution
    "E": 50,  # product
    "F": 25,  # team
    "G": 20,  # image_filename
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Enable text wrapping for content columns
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Save the workbook
wb.save("projects.xlsx")
print("Created example_data/projects.xlsx")
